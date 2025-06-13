import json
import os
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

RUTA_CLIENTES = "clientes.json"
RUTA_LOGO = "C:/Users/Usuario/Desktop/git gimnasio/Trabajo-Integrador-AJUMA-Training-Center/Gestion GImnasio Ajuma Traingin center/logo-ajuma.png"  # Ruta del logo del gimnasio  

# valores de los planes
VALORES_PLAN = {
    "mensual": 30000,
    "trimestral": 72000,  
    "anual": 216000
}

def cargar_clientes():
    """Carga la lista de clientes desde un archivo JSON."""
    if os.path.exists(RUTA_CLIENTES):
        with open(RUTA_CLIENTES, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                print("Error al decodificar JSON. Archivo corrupto o vacío.")
                return []
    return []

def exportar_factura_pdf(cliente, nombre_pdf="factura.pdf", ruta_logo=RUTA_LOGO):
    """Genera una factura en PDF para un cliente dado."""
    c = canvas.Canvas(nombre_pdf, pagesize=letter)
    ancho, alto = letter

    # logo del gym como marca de agua 
    try:
        logo = ImageReader(ruta_logo)
        c.saveState()
        c.translate(ancho / 2, alto / 2)
        c.setFillAlpha(0.1)
        logo_ancho = 300
        logo_alto = 300
        c.drawImage(logo, -logo_ancho / 2, -logo_alto / 2, width=logo_ancho, height=logo_alto, mask='auto')
        c.restoreState()
    except Exception as e:
        print(f"No se pudo cargar el logo: {e}")

    # Títulos
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(ancho / 2, alto - 100, "Factura C")
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, alto - 150, "AJUMA Training Center")
    c.setFont("Helvetica", 12)
    c.drawString(50, alto - 170, "Dirección: Formosa, Argentina")
    c.drawString(50, alto - 190, "Email: ajumatrainingcenter@gmail.com")

    # Fecha y hora
    fecha_obj = datetime.now()
    fecha_str = fecha_obj.strftime("%d/%m/%Y")
    hora_str = fecha_obj.strftime("%H:%M:%S")

    c.drawString(400, alto - 150, f"Fecha: {fecha_str}")
    c.drawString(400, alto - 170, f"Hora: {hora_str}")

    # Datos del cliente
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, alto - 230, "Datos del Cliente")
    c.setFont("Helvetica", 12)

    y = alto - 250
    campos = ["nombre", "dni", "telefono", "email", "membresia", "plan", "fecha_ingreso", "tipo_pago"]
    for campo in campos:
        valor = cliente.get(campo, "")
        c.drawString(70, y, f"{campo.capitalize().replace('_', ' ')}: {valor}")
        y -= 20

    c.save()
    print(f"Factura PDF generada: {nombre_pdf}")

def calcular_monto_mensual(plan):
    """Devuelve el monto mensual estimado según el plan."""
    plan = plan.lower()
    if plan == "mensual":
        return VALORES_PLAN["mensual"]
    elif plan == "trimestral":
        return VALORES_PLAN["trimestral"] / 3
    elif plan == "anual":
        return VALORES_PLAN["anual"] / 12
    else:
        return 0

def exportar_excel_con_grafico(clientes, nombre_archivo="clientes_ganancias.xlsx"):
    """Genera un archivo Excel con clientes, su info y gráfico de ganancia mensual."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    encabezados = [
        "Nombre", "DNI", "Teléfono", "Email", "Membresía", "Plan",
        "Método de Pago", "Fecha de Ingreso"
    ]

    anchos = [20, 15, 15, 30, 15, 12, 18, 18]
    for col, (encabezado, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    fila = 2
    ganancia_mensual_total = 0

    for cliente in clientes:
        nombre = cliente.get("nombre", "")
        dni = cliente.get("dni", "")
        telefono = cliente.get("telefono", "")
        email = cliente.get("email", "")
        membresia = cliente.get("membresia", "")
        plan_raw = cliente.get("plan", "").lower()

        if "mensual" in plan_raw:
            plan = "mensual"
        elif "trimestral" in plan_raw:
            plan = "trimestral"
        elif "anual" in plan_raw:
            plan = "anual"
        else:
            plan = ""

        pago = cliente.get("tipo_pago", "")
        fecha_ingreso = cliente.get("fecha_ingreso", "")

        monto_mensual = calcular_monto_mensual(plan)
        ganancia_mensual_total += monto_mensual

        datos_fila = [
            nombre, dni, telefono, email, membresia, plan_raw,
            pago, fecha_ingreso
        ]

        for col, valor in enumerate(datos_fila, start=1):
            ws.cell(row=fila, column=col, value=valor)
        fila += 1
        
    ws.cell(row=fila, column=7, value="Ganancia Mensual Total:")
    celda_total = ws.cell(row=fila, column=8, value=ganancia_mensual_total)
    celda_total.font = Font(bold=True)

    ws.cell(row=fila, column=7, value="Total")

    chart = BarChart()
    chart.title = "Ganancia Mensual Total"
    chart.y_axis.title = "Monto ($)"
    chart.x_axis.title = "Mes"

    data = Reference(ws, min_col=8, min_row=fila, max_row=fila)
    categories = Reference(ws, min_col=7, min_row=fila, max_row=fila)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    ws.add_chart(chart, f"A{fila + 2}")

    wb.save(nombre_archivo)
    print(f"Excel generado: {nombre_archivo}")

if __name__ == "__main__":
    clientes = cargar_clientes()

    cliente_efectivo = next((c for c in clientes if c.get("tipo_pago", "").lower() == "efectivo"), None)
    if cliente_efectivo:
        exportar_factura_pdf(cliente_efectivo, nombre_pdf="factura_cliente_efectivo.pdf")
    else:
        print("No se encontró cliente que pague en efectivo para exportar PDF.")

    exportar_excel_con_grafico(clientes, nombre_archivo="clientes_ganancias.xlsx")
